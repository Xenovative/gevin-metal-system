import json
import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from config import DEFAULT_CASH_CURRENCY, GRAMS_PER_TAEL, OUTPUT_DIR, TEMPLATE_PATH, TRANSACTION_TYPES

# 客戶單（上半部）與公司單（下半部）列號對照
# Row heights in templates/invoice_template.xlsx leave top/mid spacers so
# printed data sits in the blank frames of the A4 pre-printed 收據 form
# (branding/logo at top of each half must not be overwritten).
CUSTOMER_COPY = {
    "invoice_no_row": 4,
    "info_row": 6,
    "items_start": 11,
    "notes_row": 19,
    "total_row": 22,
    "payment_row": 23,
}
COMPANY_COPY_OFFSET = 27  # 公司單 = 客戶單列號 + 27
PAYMENT_LINE_ROWS = 4
PRINT_AREA = "A1:K54"


def _format_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if float(value) == int(value):
            return int(value)
        return round(float(value), 3)
    return value


def _parse_payment_entries(payment_method):
    if not payment_method:
        return []
    try:
        payments = json.loads(payment_method)
    except (json.JSONDecodeError, TypeError):
        return [{"method": str(payment_method), "amount": None}]
    if not isinstance(payments, list):
        return [{"method": str(payment_method), "amount": None}]
    return payments


def _format_payment_amount(amount, currency=None):
    return format_money(amount, currency)


def format_money(value, currency=None):
    """Format numeric amount with currency label, e.g. HKD$ 1,234.56."""
    if value is None or value == "":
        return None
    currency = currency or DEFAULT_CASH_CURRENCY
    num = float(value)
    if num == int(num):
        return f"{currency} {int(num):,}"
    return f"{currency} {num:,.2f}"


def format_payment_excel_lines(payment_method):
    """每種付款方式一行，供 Excel 垂直列出。"""
    lines = []
    for entry in _parse_payment_entries(payment_method):
        method = entry.get("method", "")
        amount = entry.get("amount")
        currency = entry.get("currency") or DEFAULT_CASH_CURRENCY
        if method and amount is not None:
            lines.append(f"{method} {_format_payment_amount(amount, currency)}")
        elif method:
            lines.append(method)
    return lines


def format_payment_method_display(payment_method):
    """將付款資料（JSON 或舊版純文字）轉為可讀字串。"""
    lines = format_payment_excel_lines(payment_method)
    if lines:
        return " / ".join(lines)
    if not payment_method:
        return ""
    try:
        json.loads(payment_method)
        return ""
    except (json.JSONDecodeError, TypeError):
        return str(payment_method)


def _format_amount(value, currency=None):
    return format_money(value, currency)


EXCHANGE_LABEL = "對換  (Exchange) "


def _is_company_copy(layout):
    return layout["items_start"] >= 38


def _notes_col(layout, customer_notes_col):
    """公司單備註在 E 欄；客戶單依交易類型設定。"""
    if _is_company_copy(layout):
        return 5
    return customer_notes_col


def _write_item_block(
    ws, start_row, items, has_amount=True,
    write_stock=False, source=None, destination=None,
    currency=None,
):
    """Write line items starting at start_row. Each item uses 2+ rows (gram + optional tael/oz)."""
    currency = currency or DEFAULT_CASH_CURRENCY
    row = start_row
    for item in items:
        ws.cell(row=row, column=3).value = item.get("item_type", "")
        ws.cell(row=row, column=5).value = item.get("quality", "")
        ws.cell(row=row, column=6).value = _format_number(item.get("weight_gram"))
        ws.cell(row=row, column=7).value = "克 Gram "
        if item.get("unit_price") is not None:
            ws.cell(row=row, column=9).value = format_money(item.get("unit_price"), currency)
        if item.get("amount") is not None:
            # Column K aligns with the narrow amount box on the A4 receipt form
            ws.cell(row=row, column=11).value = _format_amount(item.get("amount"), currency)
        if write_stock and source:
            ws.cell(row=row, column=8).value = f"倉存存取 {source}"

        row += 1
        if item.get("weight_tael") is not None:
            ws.cell(row=row, column=6).value = _format_number(item.get("weight_tael"))
            ws.cell(row=row, column=7).value = "両 Teal"
            if item.get("unit_price_note"):
                ws.cell(row=row, column=9).value = item.get("unit_price_note")
            if write_stock and destination:
                ws.cell(row=row, column=8).value = f"倉存位置 {destination}"
        row += 1
        if item.get("weight_oz") is not None:
            ws.cell(row=row, column=6).value = _format_number(item.get("weight_oz"))
            ws.cell(row=row, column=7).value = "安士 oz"
            row += 1
    return row


def _find_exchange_label_row(ws, search_from, search_to):
    for r in range(search_from, search_to):
        if _is_exchange_label_row(ws, r):
            return r
    return None


def _is_exchange_label_row(ws, row):
    val = ws.cell(row=row, column=3).value
    return val and "對換" in str(val)


def _clear_item_rows(ws, start_row, end_row, clear_stock=False):
    for r in range(start_row, end_row):
        if _is_exchange_label_row(ws, r):
            for c in range(4, 12):
                ws.cell(row=r, column=c).value = None
            continue
        for c in range(3, 12):
            if c == 8 and not clear_stock:
                continue
            ws.cell(row=r, column=c).value = None


def _clear_section_data(ws, layout):
    """清除範本預設資料與 XXXX 佔位符，避免公司單殘留範本內容。"""
    _clear_item_rows(
        ws, layout["items_start"], layout["notes_row"],
        clear_stock=_is_company_copy(layout),
    )

    notes_row = layout["notes_row"]
    for c in (4, 5, 10, 11):
        ws.cell(row=notes_row, column=c).value = None
    # 清除「其他」備註列（含金額欄）
    other_row = notes_row + 1
    for c in (4, 5, 10, 11):
        ws.cell(row=other_row, column=c).value = None

    ws.cell(row=layout["total_row"], column=10).value = None
    ws.cell(row=layout["total_row"], column=11).value = None
    payment_row = layout["payment_row"]
    for offset in range(PAYMENT_LINE_ROWS):
        row = payment_row + offset
        ws.cell(row=row, column=3).value = None
        # Do not clear column 5 — template already has print/goods checkboxes there
    ws.cell(row=payment_row, column=6).value = None


def _write_payment_section(ws, layout, invoice_data):
    """付款方式垂直列出；核取方塊已在 Excel 範本中，不再寫入文字。"""
    payment_row = layout["payment_row"]
    lines = format_payment_excel_lines(invoice_data.get("payment_method", ""))
    for idx, line in enumerate(lines[:PAYMENT_LINE_ROWS]):
        ws.cell(row=payment_row + idx, column=3).value = line

    handler = invoice_data.get("handler", "")
    if handler:
        ws.cell(row=payment_row, column=6).value = handler


def _fill_copy_section(
    ws, layout, invoice_data, main_items, exchange_items, tx_config, number_label, tx_date,
):
    """填入客戶單或公司單區塊。"""
    has_amount = tx_config.get("has_amount", True)
    has_exchange = tx_config.get("has_exchange", False)

    # Col K matches invoice no / date / amount boxes on the pre-printed form
    ws.cell(row=layout["invoice_no_row"], column=11).value = (
        f"{number_label} {invoice_data['invoice_no']}"
        if not str(number_label).endswith(" ")
        else f"{number_label}{invoice_data['invoice_no']}"
    )
    ws.cell(row=layout["info_row"], column=5).value = invoice_data["customer_name"]
    ws.cell(row=layout["info_row"], column=11).value = tx_date

    _clear_section_data(ws, layout)

    is_company = _is_company_copy(layout)
    if is_company:
        phone = (invoice_data.get("customer_phone") or "").strip()
        ws.cell(row=layout["info_row"], column=7).value = phone or None
    stock_kwargs = {
        "write_stock": is_company,
        "source": invoice_data.get("source_location") if is_company else None,
        "destination": invoice_data.get("destination_location") if is_company else None,
    }

    next_row = _write_item_block(
        ws, layout["items_start"], main_items, has_amount=has_amount,
        currency=invoice_data.get("invoice_currency"), **stock_kwargs,
    )

    if exchange_items and has_exchange:
        label_row = _find_exchange_label_row(
            ws, layout["items_start"], layout["notes_row"],
        )
        if label_row is None:
            label_row = min(next_row, layout["notes_row"] - 3)
            ws.cell(row=label_row, column=3).value = EXCHANGE_LABEL
        _write_item_block(
            ws, label_row + 1, exchange_items, has_amount=has_amount,
            currency=invoice_data.get("invoice_currency"), **stock_kwargs,
        )

    notes = invoice_data.get("notes", "")
    notes_row = layout["notes_row"]
    customer_notes_col = tx_config.get("customer_notes_col", 4)
    if notes:
        ws.cell(row=notes_row, column=_notes_col(layout, customer_notes_col)).value = notes

    note_amount = invoice_data.get("note_amount")
    currency = invoice_data.get("invoice_currency") or DEFAULT_CASH_CURRENCY
    if note_amount:
        ws.cell(row=notes_row, column=11).value = _format_amount(note_amount, currency)

    total = invoice_data.get("total_amount")
    if total is not None:
        ws.cell(row=layout["total_row"], column=11).value = _format_amount(total, currency)

    _write_payment_section(ws, layout, invoice_data)


def _company_layout():
    return {key: val + COMPANY_COPY_OFFSET for key, val in CUSTOMER_COPY.items()}


def _apply_a4_print_layout(ws):
    """Force A4 single-page print so Excel data aligns with pre-printed branding."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.15, right=0.15, top=0.2, bottom=0.2, header=0.0, footer=0.0
    )
    ws.print_area = PRINT_AREA

    # Do not print Excel titles over pre-printed 收據 / logo band
    for row in (3, 30):
        for col in (9, 10, 11):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            text = str(cell.value)
            if "Invoice No" in text or "編號" in text:
                continue
            if any(
                key in text
                for key in (
                    "Sales Invoice",
                    "Purchase Invoice",
                    "銷售單",
                    "購入單",
                    "兌料",
                    "交收",
                    "收據",
                    "RECEIPT",
                )
            ):
                cell.value = None


def generate_invoice_excel(invoice_data, main_items, exchange_items=None):
    """
    Generate Excel invoice from template.

    Fills both 客戶單 (top) and 公司單 (bottom, rows ~29-54) sections.
    Layout is tuned so values fall inside the blank frames of the A4
    pre-printed company receipt paper (no overlap with logo/branding).
    """
    tx_type = invoice_data["transaction_type"]
    tx_config = TRANSACTION_TYPES[tx_type]
    sheet_name = tx_config["sheet"]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{invoice_data['invoice_no']}.xlsx"
    shutil.copy(TEMPLATE_PATH, output_path)

    wb = openpyxl.load_workbook(output_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"範本中找不到分頁：{sheet_name}（現有：{', '.join(wb.sheetnames)}）"
        )

    ws = wb[sheet_name]
    tx_date = invoice_data["transaction_date"]
    if isinstance(tx_date, str):
        tx_date = datetime.strptime(tx_date, "%Y-%m-%d").date()
    elif isinstance(tx_date, datetime):
        tx_date = tx_date.date()

    number_label = tx_config["number_label"]

    _fill_copy_section(
        ws, CUSTOMER_COPY, invoice_data, main_items, exchange_items,
        tx_config, number_label, tx_date,
    )
    _fill_copy_section(
        ws, _company_layout(), invoice_data, main_items, exchange_items,
        tx_config, number_label, tx_date,
    )
    _apply_a4_print_layout(ws)

    # 只保留本次交易對應的分頁
    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]

    wb.save(output_path)
    return str(output_path)


def compute_tael_from_gram(grams):
    if grams is None or grams == "":
        return None
    return round(float(grams) / GRAMS_PER_TAEL, 3)
