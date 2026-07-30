"""Quick smoke test for create/excel/cash/review before deploy."""
from datetime import date
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from cash import build_cash_movement, signed_cash_warehouse_amount
from config import compose_receipt_storage
from invoice_generator import generate_invoice_excel, resolve_invoice_excel_path
from openpyxl import load_workbook
import database as db
import app as m


def main():
    assert signed_cash_warehouse_amount(1000, "購入") == -1000
    assert signed_cash_warehouse_amount(1000, "銷售") == 1000

    src, dst = compose_receipt_storage("存", "A倉庫")
    assert "A倉庫" in dst and "存" in dst
    src, dst = compose_receipt_storage("取", "B倉庫")
    assert "B倉庫" in src and "取" in src

    data = {
        "invoice_no": "TESTFIX_P1",
        "transaction_type": "購入",
        "customer_name": "FixTest",
        "customer_phone": "91234567",
        "transaction_date": date(2026, 7, 30),
        "handler": "Admin",
        "payment_method": "",
        "invoice_currency": "HKD$",
        "source_location": "取 客戶",
        "destination_location": "存 A倉庫",
        "notes": "n1",
        "note_amount": 0,
        "total_amount": 500.0,
        "cash_warehouse_amount": signed_cash_warehouse_amount(500, "購入"),
    }
    items = [{
        "item_type": "純銀 Silver",
        "quality": "足金",
        "weight_gram": 5,
        "weight_tael": None,
        "weight_oz": None,
        "unit_price": 100,
        "amount": 500,
    }]
    path = generate_invoice_excel(data, items)
    ws = load_workbook(path)["購入單"]
    assert ws.cell(11, 11).value == -500, ws.cell(11, 11).value
    assert ws.cell(38, 11).value == -500, ws.cell(38, 11).value
    assert ws.cell(22, 11).value == -500, ws.cell(22, 11).value

    cm = build_cash_movement(data)
    assert cm["direction"] == "out" and cm["signed_amount"] == -500

    assert resolve_invoice_excel_path(Path(path).name, "TESTFIX_P1")
    assert resolve_invoice_excel_path(path, "TESTFIX_P1")

    payload = m.load_review_page()
    assert len(payload) == 9

    text = Path(db.__file__).read_text(encoding="utf-8")
    assert text.count('"inventory_movements": [') == 1

    m.build_app()
    print("ALL SMOKE PASSED")


if __name__ == "__main__":
    main()
